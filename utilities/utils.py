import inspect
import logging
from openpyxl import Workbook,load_workbook


class Utils:
    def custom_logger(logLevel=logging.DEBUG):
        # Set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # create console handler or file handler and set the log level
        fh = logging.FileHandler("..//logs//automation.log", mode='a+')
        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                      datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger


    def read_data_from_excel(file_name,sheet):
        datalist=[]
        wb=load_workbook(filename=file_name)
        sh=wb[sheet]

        # print(sh['A3'].value)
        # print(wb['Demosheet']['A3'].value)
        # print(sh.cell(row=2,column=2).value)

        row_ct=sh.max_row
        col_ct=sh.max_column

        for i in range(1,row_ct+1):
            row=[]
            for j in range(1,col_ct+1):
                row.append(sh.cell(row=i,column=j).value)
            print('\n')
        return datalist
